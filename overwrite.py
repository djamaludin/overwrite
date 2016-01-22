# overwrite.py - Overwrite specific xlsx files.
# Copyright (C) 2015-2016, C.I.Djamaludin.
#
# Overwrite is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Overwrite is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.


import os
import datetime
import openpyxl
# Dependant on openpyxl (https://bitbucket.org/openpyxl/openpyxl) to read and write xlsx files.


#################### DEFINITIONS ####################
## Global Definitions ##
master_sheet_string = 'master_sheet' # Master sheet file needs to have the string "Master Sheet"

# Define Working Directory
working_directory = '.'
input_directory = working_directory + '/files/'
alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z'] # Don't touch
log_file = 'debug_log.txt'

# Sheet and Cell location definitions
# Issue 1 - Well Name
well_name_no_cell_loc = 'B3'
well_name_no_sheet_loc = 'Location and Ownership'
well_name_no_master_column = 'A'
well_name_no_master_column_no = alphabet.index(well_name_no_master_column) # Don't Touch!
# Issue 2a - Zone
zone_cell_loc = 'B14'
zone_sheet_loc = 'Location and Ownership'
zone_master_column = 'E'
zone_master_column_no = alphabet.index(zone_master_column) # Don't Touch!
# Issue 2b - Easting
easting_cell_loc = 'B15'
easting_sheet_loc = 'Location and Ownership'
easting_master_column = 'G'
easting_master_column_no = alphabet.index(easting_master_column) # Don't Touch!
# Issue 2c - Northing
northing_cell_loc = 'B16'
northing_sheet_loc = 'Location and Ownership'
northing_master_column = 'F'
northing_master_column_no = alphabet.index(northing_master_column) # Don't Touch!
#################### END DEFINITIONS ####################


#################### FUNCTIONS ####################
def debug_log(message):
  timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
  message_timestamp = timestamp + ',' + message
  debug_write_command = 'echo "' + message_timestamp + '" >> "'+ log_file + '"'
  os.system(debug_write_command)


def load_xlsx_file(filename):
  workbook = openpyxl.load_workbook(filename)
  # Debug Log
  log_entry = 'Okay,Load,' + str(filename)
  debug_log(log_entry)
  return workbook


def save_xlsx_file(wb, filename):
  wb.save(filename)
  # Debug Log
  log_entry = 'Okay,Save,' + str(filename)
  debug_log(log_entry)


def xlsx_cell_write(ws, cell_location, value):
  old_value = ws[cell_location].value
  ws[cell_location] = value
  # Debug Log
  log_entry = 'Okay,Overwrite,' + str(cell_location) + ',' + str(old_value) + ',' + str(value)
  debug_log(log_entry)


def generate_master_sheet_well_index(master_sheet_path):
  # Extract data from Master Sheet
  master_wb_local = load_xlsx_file(master_sheet_path)
  master_ws_local = master_wb_local.active
  # Well Index
  well_index = []
  # Extract Well column from master sheet for index
  counter_master_row = 0
  while counter_master_row < len(master_ws_local.columns[well_name_no_master_column_no]):
    well_index.append(str(master_ws_local.columns[well_name_no_master_column_no][counter_master_row].value.replace(' ','')))
    counter_master_row += 1
  return [master_wb_local, well_index]


def overwrite(filename, master_sheet_path, master_data):
  print 'Overwriting ' + str(filename)
  # Get workbook
  print 'Loading'
  wb = load_xlsx_file(filename)
  # Issue 1 - Well
  # Extract filename without extension
  filename_no_ext = os.path.splitext(os.path.basename(filename))[0]
  filename_no_ext = filename_no_ext.split('_')
  # Extract Well ID
  well_id = filename_no_ext[-1]
  # Get worksheet
  ws = wb[well_name_no_sheet_loc]
  # Overwrite Value
  print '  Well ID: ' + str(ws[well_name_no_cell_loc].value) + ' --> ' + str(well_id)
  xlsx_cell_write(ws, well_name_no_cell_loc, well_id)
  # Issue 2
  master_wb = master_data[0]
  master_ws = master_wb.active
  well_id_index = master_data[1]
  # Find corresponding Well ID
  if well_id_index.count(well_id) > 0: # Exists in Master List
    # Get index
    well_id_index_position = well_id_index.index(well_id)
    # Get new values from Master wb
    new_value_zone = master_ws.columns[zone_master_column_no][well_id_index_position].value
    new_value_easting = master_ws.columns[easting_master_column_no][well_id_index_position].value
    new_value_northing = master_ws.columns[northing_master_column_no][well_id_index_position].value
    # Write New Values
    # Issue 2a - Zone
    # Get worksheet
    ws = wb[zone_sheet_loc]
    # Overwrite Value
    print '  Zone: ' + str(ws[zone_cell_loc].value) + ' --> ' + str(new_value_zone)
    xlsx_cell_write(ws, zone_cell_loc, new_value_zone)
    # Issue 2b - Easting
    # Get worksheet
    ws = wb[easting_sheet_loc]
    # Overwrite Value
    print '  Easting: ' + str(ws[easting_cell_loc].value) + ' --> ' + str(new_value_easting)
    xlsx_cell_write(ws, easting_cell_loc, new_value_easting)
    # Issue 2c - Northing
    # Get worksheet
    ws = wb[northing_sheet_loc]
    # Overwrite Value
    print '  Northing: ' + str(ws[northing_cell_loc].value) + ' --> ' + str(new_value_northing)
    xlsx_cell_write(ws, northing_cell_loc, new_value_northing)
  else:
    # Debug Log
    log_entry = 'Error,Well ID Missing in Master List,' + str(well_id) + ',' + str(master_sheet_path)
    debug_log(log_entry)
  # Save xlsx file
  print 'Saving'
  save_xlsx_file(wb, filename)
  print ''
#################### END FUNCTIONS ####################


#################### MAIN ####################
# Header
os.system('cls')
print ''
print 'Overwrite (version 1.0)'
print '======================='
print 'Unrestricted bulk manipulation of data'
print 'View overwrite.py for options.'
print ''
raw_input('Press any key to continue with Overwrite operation.')
# Get input directory files
input_files = os.listdir(input_directory)
# Check for open files, you can't concurrently write to the same file - Sync issues!
counter_input_files = 0
open_file_count = 0
open_file_list = []
master_sheet_list = []
while counter_input_files < len(input_files):
  if input_files[counter_input_files].count('~$') > 0: # There are files open!
    open_file_count += 1
    open_file_list.append(str(input_directory + input_files[counter_input_files]).replace('~$',''))
    # Debug Log
    log_entry = 'Error,XLSX File Open (File Sync Issue),' + str(input_directory + input_files[counter_input_files])
    debug_log(log_entry)
  else:
    pass
  potential_master_sheet = input_files[counter_input_files].lower().replace(' ','_')
  if potential_master_sheet.count(master_sheet_string) > 0: # filename is potential master sheet
    master_sheet_list.append(input_files[counter_input_files])
  else:
    pass
  counter_input_files += 1
if open_file_count == 0: # Files are open, close
  # Check if a master sheet exists
  if len(master_sheet_list) > 1: # Greater than 1 master sheet exsists - This is confusing
    # Error and exit
    # Debug Log
    log_entry = 'Error,Multiple Master Sheets,Need to clarify which one'
    debug_log(log_entry)
    print 'Error: Duplicate Master Sheets in directory.'
  elif len(master_sheet_list) < 1: # Less than 1 master sheet exsists - This is bad
    # Error and exit
    # Debug Log
    log_entry = 'Error,Master Sheet Missing'
    debug_log(log_entry)
    print 'Error: No Master Sheet in directory.'
  else:  # Exactly 1 master sheet exsists - This is good
    master_sheet = master_sheet_list[0]
    # Grab Master Sheet data
    master_sheet_path = input_directory + master_sheet
    master_data = generate_master_sheet_well_index(master_sheet_path)
    # Remove master_sheet from list of file operations
    master_sheet_index = input_files.index(master_sheet)
    del input_files[master_sheet_index]
    # Do Issue 1 operation for each file
    counter_file = 0
    while counter_file < len(input_files):
      filename = input_directory + input_files[counter_file]
      overwrite(filename, master_sheet_path, master_data)
      counter_file += 1
# Old Code
#  else: # No Master Sheet
#    # Error and exit
#    # Debug Log
#    log_entry = 'Error,Master Sheet Missing,' + str(master_sheet_path)
#    debug_log(log_entry)
else: # Files are open, close before progressing. Bad idea to continue, file sync issues.
  print 'Error: Please close the following files:'
  print ''
  counter_open_file_list = 0
  while counter_open_file_list < len(open_file_list):
    print str(open_file_list[counter_open_file_list])
    counter_open_file_list += 1
# Finish
print ''
print 'Check log file for Overwrite operations: ' + str(input_directory + log_file)
raw_input('Press any key to exit...')
#################### END MAIN ####################
